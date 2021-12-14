import subprocess
import re
import random
import base64

import willump
import requests
import openpyxl
import PySimpleGUI as sg

from lcu import LCU


def get_challenge_list(sheet, role):
    col_dict = {'top': 1, 'jungle': 2, 'mid': 3, 'bot': 4, 'support': 5}
    col = col_dict[role]
    role_list = []
    i = 2
    while sheet.cell(i, col).value:
        role_list.append(sheet.cell(i, col).value)
        i += 1
    return role_list


# def get_client_port_and_auth():
#     out = subprocess.Popen("wmic PROCESS WHERE name='LeagueClientUx.exe' GET commandline",
#                            stdout=subprocess.PIPE, stderr=subprocess.PIPE).communicate()
#     port = re.findall('--app-port=([0-9]*)', str(out))[0]
#     auth = re.findall(r'--remoting-auth-token=([\w-]*)', str(out))[0]
#     return port, auth
#
# def get_summoner_name():
#     port, password = get_client_port_and_auth()
#
#     url = f'https://127.0.0.1:{port}/lol-summoner/v1/current-summoner'
#     print(url)
#     print(password)
#     auth = base64.b64encode(password.encode('utf-8')).decode('utf-8')
#     headers = {'Authorization': f'Basic {auth}'}
#     print(headers)
#
#     with requests.Session() as s:
#         response = requests.get(url, headers=headers, verify='riotgames.pem')
#     print(response)


def get_summoner_name():
    lcu = LCU()
    return lcu.summoner_info['internalName']


def main():
    result = get_summoner_name()
    print(result)
    return
    expenses_excel = openpyxl.load_workbook("challenges_roles.xlsx")
    curr_sheet = expenses_excel['challenges']
    all_challenges = {
        'Top': get_challenge_list(curr_sheet, 'top'),
        'Jungle': get_challenge_list(curr_sheet, 'jungle'),
        'Mid': get_challenge_list(curr_sheet, 'mid'),
        'Bot': get_challenge_list(curr_sheet, 'bot'),
        'Support': get_challenge_list(curr_sheet, 'support'),
    }

    categories = ['hello']
    charges = []

    # Setting up the Gui
    sg.ChangeLookAndFeel('BlueMono')

    layout = [
        [sg.Text('Role'), sg.Combo(
            ['Top', 'Jungle', 'Mid', 'Bot', 'Support'], key='role')],
        [sg.Submit(), sg.Exit()]
    ]

    window = sg.Window("LOL Challenges", layout, size=(250,80), element_justification='c')

    while True:
        button, values = window.Read()
        if button is None or button == 'Exit':
            break
        # Adding the values from the gui into a list of dictionaries for each time the form was submitted
        if button == 'Submit':
            sg.Popup(random.choice(all_challenges[values['role']]))

    window.Close()

    print(charges)


if __name__ == "__main__":
    main()
